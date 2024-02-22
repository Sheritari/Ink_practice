from openpyxl import Workbook
from django.http import HttpResponse
from json import dumps
from .models import Characteristic, CharacteristicType, Well, WellCharacteristicBinding
from os.path import join
from django.conf import settings

def export_characteristics_to_excel():
    file_path = join(settings.MEDIA_ROOT, 'characteristics.xlsx')  # Путь сохранения файла на сервере

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="characteristics.xlsx"'

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = 'Characteristics'

    # Названия колонок в Excel
    column_names = [
        'id',
        Characteristic._meta.get_field('name').verbose_name,
        Characteristic._meta.get_field('characteristic_type').verbose_name,
        Characteristic._meta.get_field('value').verbose_name,
    ]
    worksheet.append(column_names)

    # Запись данных модели в Excel
    for characteristic in Characteristic.objects.all():
        value_str = dumps(characteristic.value)  # Преобразование JSON-объекта в строку
        row_data = [
            characteristic.id,
            getattr(characteristic, 'name'),
            characteristic.characteristic_type.name,
            value_str
        ]
        worksheet.append(row_data)

    string_columns = [2, 3]  # Номера столбцов в которых нужно установить тип данных как строку

    for col_num in string_columns:
        for cell in worksheet.iter_cols(min_col=col_num, max_col=col_num):
            for c in cell:
                c.number_format = '@'


    workbook.save(file_path)
    return file_path

def export_data_to_excel():
    workbook = Workbook()

    worksheet1 = workbook.active
    worksheet1.title = 'Characteristic Types'

    column_names = [
        'id',
        CharacteristicType._meta.get_field('name').verbose_name,
        CharacteristicType._meta.get_field('max_value_x').verbose_name,
        CharacteristicType._meta.get_field('max_value_y').verbose_name,
        CharacteristicType._meta.get_field('json_key_name').verbose_name,
        CharacteristicType._meta.get_field('json_value_name').verbose_name,
    ]
    worksheet1.append(column_names)

    for characteristictype in CharacteristicType.objects.all():
        row_data = [
            characteristictype.id,
            getattr(characteristictype, 'name'),
            getattr(characteristictype, 'max_value_x'),
            getattr(characteristictype, 'max_value_y'),
            getattr(characteristictype, 'json_key_name'),
            getattr(characteristictype, 'json_value_name')
        ]
        worksheet1.append(row_data)

    string_columns = [2, 5, 6]  # Номера столбцов в которых нужно установить тип данных как строку

    for col_num in string_columns:
        for cell in worksheet1.iter_cols(min_col=col_num, max_col=col_num):
            for c in cell:
                c.number_format = '@'

    worksheet2 = workbook.create_sheet(title='Characteristics')

    column_names = [
        'id',
        Characteristic._meta.get_field('name').verbose_name,
        Characteristic._meta.get_field('characteristic_type').verbose_name,
        Characteristic._meta.get_field('value').verbose_name,
    ]
    worksheet2.append(column_names)

    for characteristic in Characteristic.objects.all():
        value_str = dumps(characteristic.value)  # Преобразование JSON-объекта в строку
        row_data = [
            characteristic.id,
            getattr(characteristic, 'name'),
            characteristic.characteristic_type.name,
            value_str
        ]
        worksheet2.append(row_data)

    string_columns = [2, 3]  # Номера столбцов в которых нужно установить тип данных как строку

    for col_num in string_columns:
        for cell in worksheet2.iter_cols(min_col=col_num, max_col=col_num):
            for c in cell:
                c.number_format = '@'

    worksheet3 = workbook.create_sheet(title='Wells')

    column_names = [
        'id',
        Well._meta.get_field('name').verbose_name,
    ]
    worksheet3.append(column_names)

    for well in Well.objects.all():
        row_data = [
            well.id,
            getattr(well, 'name'),
        ]
        worksheet3.append(row_data)

    string_columns = [2, 3]  # Номера столбцов в которых нужно установить тип данных как строку

    for col_num in string_columns:
        for cell in worksheet3.iter_cols(min_col=col_num, max_col=col_num):
            for c in cell:
                c.number_format = '@'

    worksheet4 = workbook.create_sheet(title='WellCharacteristicBindings')

    column_names = [
        'id',
        WellCharacteristicBinding._meta.get_field('well').verbose_name,
        WellCharacteristicBinding._meta.get_field('characteristic').verbose_name
    ]
    worksheet4.append(column_names)

    for wellchar in WellCharacteristicBinding.objects.all():
        row_data = [
            wellchar.id,
            wellchar.well.name,
            wellchar.characteristic.name
        ]
        worksheet4.append(row_data)

    string_columns = [2, 3]  # Номера столбцов в которых нужно установить тип данных как строку

    for col_num in string_columns:
        for cell in worksheet4.iter_cols(min_col=col_num, max_col=col_num):
            for c in cell:
                c.number_format = '@'

    file_path = 'exported_data.xlsx'
    workbook.save(file_path)
    return file_path